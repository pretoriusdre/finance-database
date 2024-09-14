from openpyxl import load_workbook
import pandas as pd


class ExcelFile():

    def __init__(self, input_file):

        self.input_file = input_file
        wb = load_workbook(self.input_file, data_only=True)
        self.tables = {}
        self.metadata = {}
        for ws in wb.worksheets:
            for entry, data_boundary in ws.tables.items():
                # Parse the data within the ref boundary
                data_region = ws[data_boundary]
                # The inner list comprehension gets the values for each cell in the table
                content = [[cell.value for cell in row] for row in data_region]
                header = content[0]
                rest = content[1:]
                df = pd.DataFrame(rest, columns=header)
                df = df.astype(object).where(df.notnull(), None)
                self.tables[entry] = df

        metadata = self.tables.get('metadata')
    
        if metadata:
            for ind, record in metadata_df.iterrows():
                self.metadata.setdefault(list(record)[0], list(record)[1])





